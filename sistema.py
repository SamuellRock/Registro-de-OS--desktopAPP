#-*- coding:utf-8 -*-
from typing import Tuple
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook
from tkcalendar import Calendar


# Função para chamar as Listas dos Arquivos .txt
def lista_bairros():
    with open('Lista_de_Bairros.txt', 'r') as arquivo:
        texto = arquivo.read()
    listaTexto = texto.split("\n")
    return listaTexto

def tipo_OS():
    with open('Lista_de_OS.txt', 'r') as arquivo:
        tipoOS = arquivo.read()
    listaOS = tipoOS.split("\n")
    return listaOS


# Aparencia padrao do sistema
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")



class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        #self.aparencia()
        self.todoSistema()

    def layout_config(self):
        self.title("Registro de Ordem de Serviço")
        self.geometry("700x500")
       

   # def aparencia(self):
        #self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=["#000", '#fff']).place(x=50, y=430)
        #self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50, y=460)

    

    def todoSistema(self):
        # Barra no Cabeçalho ------------------------------------------
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="#d5cd1a", fg_color="#d5cd1a")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Registro de Ordem de Serviço", font=("Century Gothic Negrito", 24), text_color="#000").place(x=190, y=10)

        span = ctk.CTkLabel(self, text="Por favor, Preencha Todos os campos", font=("Century Gothic Negrito", 16), text_color=["#000","#fff"]).place(x=50, y=70)



        #CRIAR PLANILHA .xlsx -----------------------------------------------
        ficheiro = pathlib.Path('REGISTRO-de-OS.xlsx')

        if ficheiro.exists():
            pass
        else:
            ficheiro = Workbook()
            folha=ficheiro.active
            folha['A1']='Data'
            folha['B1']='Nome do Operador'
            folha['C1']='ID OS'
            folha['D1']='Localida/Bairro'
            folha['E1']='Alteração'
            folha['F1']='Observações sobre a OS'

            ficheiro.save("REGISTRO-de-OS.xlsx")



        # Salvar dados na Planilha -----------------------------------------
        def submit():

            data = dataValue.get()
            name = nameValue.get()
            numero = numeroValue.get()
            local = formLocal.get()
            tipo = formtipoOS.get()
            obs = formObs.get(0.0, END)

            if (data=="" or name=="" or numero==""):
                messagebox.showerror("Sistema", "ERRO!\nPor favor, preencha todos os campos.")
            else:
                ficheiro = openpyxl.load_workbook('REGISTRO-de-OS.xlsx')
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row+1, value=data)
                folha.cell(column=2, row=folha.max_row, value=name)
                folha.cell(column=3, row=folha.max_row, value=numero)
                folha.cell(column=4, row=folha.max_row, value=local)
                folha.cell(column=5, row=folha.max_row, value=tipo)
                folha.cell(column=6, row=folha.max_row, value=obs)

                ficheiro.save(r"REGISTRO-de-OS.xlsx")
                messagebox.showinfo("Sistema", "Dados Salvos")
                clear()


        # Função de Limpar todos os Campos -----------------------------------
        def clear():
            data = dataValue.set("")
            name = nameValue.set("")
            numero = numeroValue.set("")
            obs = formObs.delete(0.0, END)



        # tranformar os valores dos camps em tipo string
        dataValue = StringVar()
        nameValue = StringVar()
        numeroValue = StringVar()



        #Input
        formName = ctk.CTkEntry(self, width=350, textvariable=nameValue, font=("Century Gothic Bold", 16), fg_color="transparent")
        formData = ctk.CTkEntry(self, width=150, textvariable=dataValue, font=("Century Gothic Bold", 16), fg_color="transparent")
        formNumero = ctk.CTkEntry(self, width=150, textvariable=numeroValue, font=("Century Gothic", 16), fg_color="transparent")

        #Combo box 
        formLocal = ctk.CTkComboBox(self, values=lista_bairros(), font=("century gothic", 14))
        formtipoOS = ctk.CTkComboBox(self, values=tipo_OS(), font=("century gothic", 14), width=150)
        formtipoOS.set("Vazamento")

        #Caixa Observação
        formObs = ctk.CTkTextbox(self, width=460, height=140, font=("arial", 14), border_color="#aaa", border_width=2, fg_color="transparent")

        #Labels
        lb_formName = ctk.CTkLabel(self, text="Operador:", font=("Century Gothic Negrito", 16), text_color=["#000","#fff"])
        lb_formData = ctk.CTkLabel(self, text="Data:", font=("Century Gothic Negrito", 16), text_color=["#000","#fff"])
        lb_formLocal = ctk.CTkLabel(self, text="Localidade:", font=("Century Gothic Negrito", 16), text_color=["#000","#fff"])
        #lb_formNumero = ctk.CTkLabel(self, text="Numero da OS:", font=("Century Gothic Negrito", 16), text_color=["#000","#fff"])
        lb_formTipoOS = ctk.CTkLabel(self, text="Tipo de OS:", font=("Century Gothic Negrito", 16), text_color=["#000","#fff"])
        lb_formObs = ctk.CTkLabel(self, text="Observações:", font=("Century Gothic Negrito", 16), text_color=["#000","#fff"])

        #Buttons
        btn_submit = ctk.CTkButton(self, text="Salvar Dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=420)
        btn_clear = ctk.CTkButton(self, text="Limpar Campos".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=420)


        #Posição dos Elemntos
        lb_formName.place(x=50, y=120)
        formName.place(x=50, y=150)

        lb_formData.place(x=500,y=120)
        formData.place(x=500, y=150)

        #lb_formNumero.place(x=300, y=190)
        formNumero.place(x=300, y=220)

        lb_formTipoOS.place(x=500, y=190)
        formtipoOS.place(x=500, y=220)

        lb_formLocal.place(x=50, y=190)
        formLocal.place(x=50, y=220)

        lb_formObs.place(x=50, y=260)
        formObs.place(x=185, y=270)
         


    # modulo para mudar o tema da aparencia ***
    #def change_apm(self, nova_aparencia):
        #ctk.set_appearance_mode(nova_aparencia)


if __name__=="__main__":
    app = App()
    app.mainloop()